import argparse
import config
import RELEVEN.PBWstarConstants
from openpyxl import load_workbook
from rdflib import Literal, RDF, RDFS
from rdflib.plugins.stores import sparqlstore

# a bunch of constants
HEADROW = 2

def save_headers(ws):
    headlist = dict()
    for i, col in enumerate(ws[HEADROW]):
        headlist[col.value] = i+1
    return headlist



if __name__ == '__main__':
    # Parse the arguments
    parser = argparse.ArgumentParser(
        prog="clean_locations",
        description="Modify the old PBW location factoids according to data gathered in a spreadsheet."
    )
    parser.add_argument('-f', '--file',
                        required=True,
                        help="Spreadsheet file with location information")
    parser.add_argument('-r', '--reader',
                        required=True,
                        help="URI for the RELEVEN team member who did this")
    args = parser.parse_args()

    # Get the connection to the triple store
    store = sparqlstore.SPARQLUpdateStore(config.graphuri, config.graphuri + '/statements', method='POST',
                                          auth=(config.graphuser, config.graphpw))
    c = RELEVEN.PBWstarConstants.PBWstarConstants(store=store)
    g = c.graph

    # Parse the spreadsheet
    wb = load_workbook(filename=args.file, read_only=True)
    ws = wb.active
    headers = save_headers(ws)
    for row in ws.iter_rows(min_row=HEADROW+1, values_only=True):
        # Define our lookup function
        def gv(header):
            return row[headers[header]]

        if gv('RELEVEN location type') == 'Specifies presence':
            # Find the event and upgrade it to an activity
            locevent = g.value(None, RDFS.label, f"Location event for factoid {gv(row, 'Factoid ID')}")
            g.add((locevent, RDF.type, c.entitylabels['E7']))
            # Replace the label with the label from the factoid
            g.remove((locevent, RDFS.label, None))
            g.add((locevent, RDFS.label, Literal(gv('Description'), lang="en")))

            # Is it dated?
            if gv('Date suggested'):
                pass
                # ...but this reading was done by the RELEVEN member


